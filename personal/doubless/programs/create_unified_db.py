#!/usr/bin/env python3
"""
통합 급여 관리 DB 생성 스크립트

통합 DB 구조:
1. employees - 직원 마스터 (개인정보)
2. salary_records - 수업내역 로우데이터 (트레이너별 회원별 수업 기록)
3. trainer_monthly_salary - 트레이너 월별 기본급/인센티브
4. info_staff_salary - 인포 직원 급여

데이터 소스:
- salary.db: salary_records (수업내역)
- payroll.db: employees, info_staff_salary
- xlsx: 트레이너 기본급/인센티브 정보
"""

import sqlite3
from pathlib import Path
import openpyxl
import shutil
from datetime import datetime

# 경로 설정
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / 'data'
EXCEL_PATH = BASE_DIR / 'pay_result' / 'excel_data' / '목포급여_20261006.xlsx'

OLD_SALARY_DB = DATA_DIR / 'salary.db'
OLD_PAYROLL_DB = DATA_DIR / 'payroll.db'
UNIFIED_DB = DATA_DIR / 'unified_salary.db'


def create_unified_db():
    """통합 DB 스키마 생성"""
    # 기존 파일 백업
    if UNIFIED_DB.exists():
        backup_path = DATA_DIR / f'backups/unified_salary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        backup_path.parent.mkdir(exist_ok=True)
        shutil.copy(UNIFIED_DB, backup_path)
        print(f"✅ 기존 DB 백업: {backup_path}")
        UNIFIED_DB.unlink()

    conn = sqlite3.connect(UNIFIED_DB)
    cursor = conn.cursor()

    # 1. 직원 테이블 (마스터)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            job_type TEXT NOT NULL,  -- '트레이너' or '인포'
            bank TEXT,
            account_number TEXT,
            resident_number TEXT,
            status TEXT DEFAULT '근무',  -- '근무' or '퇴사'
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(name, resident_number)
        )
    ''')

    # 2. 수업내역 테이블 (로우데이터) - salary.db에서 가져옴
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS salary_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            년도 INTEGER NOT NULL,
            월 TEXT NOT NULL,
            트레이너 TEXT NOT NULL,
            NO INTEGER,
            회원명 TEXT NOT NULL,
            성별 TEXT,
            등록세션 REAL,
            총진행세션 REAL,
            남은세션 REAL,
            결제형태 TEXT,
            등록비용 REAL,
            공급가 REAL,
            회단가 REAL,
            매출대비율 REAL,
            수업료_정산 REAL,
            당월진행세션 REAL,
            당월수업료 REAL,
            이달의매출 REAL,
            등록일시 TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(년도, 월, 트레이너, 회원명)
        )
    ''')

    # 3. 트레이너 월별 급여 테이블 (기본급/인센티브)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trainer_monthly_salary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            trainer_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            base_salary REAL DEFAULT 0,  -- 기본급
            incentive REAL DEFAULT 0,    -- 인센티브
            tuition_fee REAL DEFAULT 0,  -- 수업료 (salary_records 합계와 비교용)
            base_incentive_payment_date DATE,  -- 기본급+인센 지급일
            tuition_payment_date DATE,         -- 수업료 지급일
            total_salary REAL,           -- 총급여 (기본급+인센+수업료)
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES employees(id),
            UNIQUE(trainer_name, year, month)
        )
    ''')

    # 4. 인포 직원 급여 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS info_staff_salary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            staff_name TEXT NOT NULL,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            base_salary REAL DEFAULT 0,
            salary_after_tax REAL,
            payment_date DATE,
            extra_pay REAL DEFAULT 0,
            extra_pay_note TEXT,
            total_salary REAL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES employees(id),
            UNIQUE(staff_name, year, month)
        )
    ''')

    # 인덱스 생성
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_salary_trainer ON salary_records(트레이너)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_salary_month ON salary_records(년도, 월)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_trainer_salary_month ON trainer_monthly_salary(trainer_name, year, month)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_info_salary_month ON info_staff_salary(staff_name, year, month)')

    conn.commit()
    conn.close()
    print(f"✅ 통합 DB 스키마 생성 완료: {UNIFIED_DB}")


def migrate_employees():
    """payroll.db에서 직원 정보 마이그레이션"""
    if not OLD_PAYROLL_DB.exists():
        print("⚠️ payroll.db 없음, 직원 정보 스킵")
        return

    src_conn = sqlite3.connect(OLD_PAYROLL_DB)
    dst_conn = sqlite3.connect(UNIFIED_DB)

    src_cur = src_conn.cursor()
    dst_cur = dst_conn.cursor()

    src_cur.execute('SELECT name, job_type, bank, account_number, resident_number, status FROM employees')
    rows = src_cur.fetchall()

    count = 0
    for row in rows:
        try:
            dst_cur.execute('''
                INSERT OR REPLACE INTO employees
                (name, job_type, bank, account_number, resident_number, status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', row)
            count += 1
        except Exception as e:
            print(f"  ⚠️ 직원 입력 실패: {row[0]} - {e}")

    dst_conn.commit()
    src_conn.close()
    dst_conn.close()
    print(f"✅ 직원 {count}명 마이그레이션 완료")


def migrate_salary_records():
    """salary.db에서 수업내역 마이그레이션"""
    if not OLD_SALARY_DB.exists():
        print("⚠️ salary.db 없음, 수업내역 스킵")
        return

    src_conn = sqlite3.connect(OLD_SALARY_DB)
    dst_conn = sqlite3.connect(UNIFIED_DB)

    src_cur = src_conn.cursor()
    dst_cur = dst_conn.cursor()

    src_cur.execute('SELECT * FROM salary_records')
    rows = src_cur.fetchall()

    # 컬럼명 가져오기
    src_cur.execute('PRAGMA table_info(salary_records)')
    columns = [col[1] for col in src_cur.fetchall()]

    count = 0
    for row in rows:
        try:
            # id 제외하고 삽입
            values = row[1:]  # id 제외
            cols = columns[1:]  # id 제외
            placeholders = ','.join(['?' for _ in cols])
            col_names = ','.join(cols)

            dst_cur.execute(f'''
                INSERT OR REPLACE INTO salary_records ({col_names})
                VALUES ({placeholders})
            ''', values)
            count += 1
        except Exception as e:
            print(f"  ⚠️ 수업내역 입력 실패: {row} - {e}")

    dst_conn.commit()
    src_conn.close()
    dst_conn.close()
    print(f"✅ 수업내역 {count}건 마이그레이션 완료")


def load_trainer_salary_from_xlsx():
    """xlsx에서 트레이너 기본급/인센티브 로드"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['트레이너']

    conn = sqlite3.connect(UNIFIED_DB)
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or str(name).strip() == '' or str(name).strip() == '이름':
            continue

        month_str = str(row[2]).replace('월', '').strip() if row[2] else ''
        if not month_str.isdigit():
            continue

        # 데이터 파싱
        base_salary = float(row[3]) if row[3] else 0
        incentive = float(row[4]) if row[4] else 0
        tuition_fee = float(row[7]) if row[7] else 0
        total_salary = float(row[10]) if row[10] else 0

        # 지급일 파싱
        base_payment_date = None
        tuition_payment_date = None
        if row[6]:
            try:
                base_payment_date = row[6].strftime('%Y-%m-%d') if hasattr(row[6], 'strftime') else str(row[6]).split()[0]
            except:
                pass
        if row[9]:
            try:
                tuition_payment_date = row[9].strftime('%Y-%m-%d') if hasattr(row[9], 'strftime') else str(row[9]).split()[0]
            except:
                pass

        notes = str(row[15]).strip() if row[15] else ''

        # employee_id 조회
        cursor.execute('SELECT id FROM employees WHERE name = ?', (str(name).strip(),))
        emp_result = cursor.fetchone()
        emp_id = emp_result[0] if emp_result else None

        try:
            cursor.execute('''
                INSERT OR REPLACE INTO trainer_monthly_salary
                (employee_id, trainer_name, year, month, base_salary, incentive,
                 tuition_fee, base_incentive_payment_date, tuition_payment_date,
                 total_salary, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, str(name).strip(), 2025, int(month_str), base_salary,
                  incentive, tuition_fee, base_payment_date, tuition_payment_date,
                  total_salary, notes))
            count += 1
        except Exception as e:
            print(f"  ⚠️ 트레이너 급여 입력 실패: {name} {month_str}월 - {e}")

    conn.commit()
    conn.close()
    wb.close()
    print(f"✅ 트레이너 월별 급여 {count}건 입력 완료")


def load_info_salary_from_xlsx():
    """xlsx에서 인포 직원 급여 로드"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['직원(인포)']

    conn = sqlite3.connect(UNIFIED_DB)
    cursor = conn.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or str(name).strip() == '' or str(name).strip() == '이름':
            continue

        month_str = str(row[2]).replace('월', '').strip() if row[2] else ''
        if not month_str.isdigit():
            continue

        base_salary = float(row[3]) if row[3] else 0
        salary_after_tax = float(row[4]) if row[4] else 0

        payment_date = None
        if row[5]:
            try:
                payment_date = row[5].strftime('%Y-%m-%d') if hasattr(row[5], 'strftime') else str(row[5]).split()[0]
            except:
                pass

        extra_pay = float(row[6]) if row[6] else 0
        extra_pay_note = str(row[7]).strip() if row[7] else ''
        total_salary = float(row[8]) if row[8] else 0

        # employee_id 조회
        cursor.execute('SELECT id FROM employees WHERE name = ?', (str(name).strip(),))
        emp_result = cursor.fetchone()
        emp_id = emp_result[0] if emp_result else None

        try:
            cursor.execute('''
                INSERT OR REPLACE INTO info_staff_salary
                (employee_id, staff_name, year, month, base_salary, salary_after_tax,
                 payment_date, extra_pay, extra_pay_note, total_salary)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, str(name).strip(), 2025, int(month_str), base_salary,
                  salary_after_tax, payment_date, extra_pay, extra_pay_note, total_salary))
            count += 1
        except Exception as e:
            print(f"  ⚠️ 인포 급여 입력 실패: {name} {month_str}월 - {e}")

    conn.commit()
    conn.close()
    wb.close()
    print(f"✅ 인포 월별 급여 {count}건 입력 완료")


def show_summary():
    """통합 DB 요약"""
    conn = sqlite3.connect(UNIFIED_DB)
    cursor = conn.cursor()

    print("\n" + "="*60)
    print("📊 통합 DB 요약")
    print("="*60)

    # 직원 수
    cursor.execute('SELECT job_type, status, COUNT(*) FROM employees GROUP BY job_type, status')
    print("\n👥 직원 현황:")
    for row in cursor.fetchall():
        print(f"  - {row[0]} ({row[1]}): {row[2]}명")

    # 수업내역 (salary_records)
    cursor.execute('''
        SELECT 트레이너, COUNT(DISTINCT 월) as 월수, COUNT(*) as 레코드수,
               SUM(당월수업료) as 수업료합계
        FROM salary_records
        GROUP BY 트레이너
    ''')
    print("\n📝 수업내역 (salary_records):")
    for row in cursor.fetchall():
        print(f"  - {row[0]}: {row[1]}개월, {row[2]}건, 수업료 {row[3]:,.0f}원")

    # 트레이너 월별 급여
    cursor.execute('''
        SELECT trainer_name, COUNT(*) as 월수,
               SUM(base_salary) as 기본급합계,
               SUM(incentive) as 인센합계,
               SUM(tuition_fee) as 수업료합계
        FROM trainer_monthly_salary
        GROUP BY trainer_name
    ''')
    print("\n💰 트레이너 월별 급여 (trainer_monthly_salary):")
    for row in cursor.fetchall():
        print(f"  - {row[0]}: {row[1]}개월, 기본급 {row[2]:,.0f}, 인센 {row[3]:,.0f}, 수업료 {row[4]:,.0f}")

    # 인포 급여
    cursor.execute('''
        SELECT staff_name, COUNT(*) as 월수, SUM(total_salary) as 총급여
        FROM info_staff_salary
        GROUP BY staff_name
    ''')
    print("\n💰 인포 급여 (info_staff_salary):")
    for row in cursor.fetchall():
        total = row[2] if row[2] else 0
        print(f"  - {row[0]}: {row[1]}개월, 총 {total:,.0f}원")

    conn.close()


def verify_tuition():
    """수업료 검증 (salary_records vs trainer_monthly_salary)"""
    conn = sqlite3.connect(UNIFIED_DB)
    cursor = conn.cursor()

    print("\n" + "="*60)
    print("🔍 수업료 검증 (salary_records vs trainer_monthly_salary)")
    print("="*60)

    # salary_records 집계
    cursor.execute('''
        SELECT 트레이너, 월, SUM(당월수업료) as 수업료
        FROM salary_records
        GROUP BY 트레이너, 월
    ''')
    records_data = {(row[0], row[1]): row[2] for row in cursor.fetchall()}

    # trainer_monthly_salary 조회
    cursor.execute('''
        SELECT trainer_name, month||'월', tuition_fee
        FROM trainer_monthly_salary
    ''')
    salary_data = {(row[0], row[1]): row[2] for row in cursor.fetchall()}

    all_keys = set(records_data.keys()) | set(salary_data.keys())

    issues = []
    for key in sorted(all_keys):
        rec = records_data.get(key, 0) or 0
        sal = salary_data.get(key, 0) or 0
        diff = sal - rec

        if abs(diff) >= 10:
            issues.append((key[0], key[1], rec, sal, diff))

    if issues:
        print(f"\n⚠️ 불일치 {len(issues)}건:")
        print(f"\n{'트레이너':<8} {'월':<5} {'records':>14} {'salary':>14} {'차이':>14}")
        print("-"*60)
        for item in issues:
            print(f"{item[0]:<8} {item[1]:<5} {item[2]:>14,.0f} {item[3]:>14,.0f} {item[4]:>+14,.0f}")
    else:
        print("\n✅ 모든 수업료 일치!")

    conn.close()


def main():
    print("="*60)
    print("🚀 통합 급여 관리 DB 생성")
    print("="*60)
    print(f"📁 엑셀: {EXCEL_PATH}")
    print(f"📁 기존 salary.db: {OLD_SALARY_DB}")
    print(f"📁 기존 payroll.db: {OLD_PAYROLL_DB}")
    print(f"📁 통합 DB: {UNIFIED_DB}")

    # 1. DB 스키마 생성
    print("\n[1/5] DB 스키마 생성...")
    create_unified_db()

    # 2. 직원 정보 마이그레이션
    print("\n[2/5] 직원 정보 마이그레이션...")
    migrate_employees()

    # 3. 수업내역 마이그레이션
    print("\n[3/5] 수업내역 마이그레이션...")
    migrate_salary_records()

    # 4. 트레이너 급여 로드
    print("\n[4/5] 트레이너 월별 급여 로드...")
    load_trainer_salary_from_xlsx()

    # 5. 인포 급여 로드
    print("\n[5/5] 인포 월별 급여 로드...")
    load_info_salary_from_xlsx()

    # 요약
    show_summary()

    # 검증
    verify_tuition()

    print("\n" + "="*60)
    print("✅ 통합 DB 생성 완료!")
    print("="*60)


if __name__ == '__main__':
    main()
