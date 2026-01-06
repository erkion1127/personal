#!/usr/bin/env python3
"""
급여 관리 DB 생성 스크립트
- 직원 정보 (employees)
- 트레이너 급여 (trainer_salary)
- 인포 직원 급여 (info_staff_salary)
"""

import sqlite3
from pathlib import Path
import openpyxl
from datetime import datetime

# 경로 설정
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / 'data'
EXCEL_PATH = BASE_DIR / 'pay_result' / 'excel_data' / '목포급여_20261006.xlsx'
DB_PATH = DATA_DIR / 'payroll.db'


def create_database():
    """DB 스키마 생성"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 직원 테이블 (마스터)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            job_type TEXT NOT NULL,  -- '트레이너' or '인포'
            bank TEXT,
            account_number TEXT,
            resident_number TEXT,
            status TEXT DEFAULT '근무',  -- '근무' or '퇴사'
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(name, resident_number)
        )
    ''')

    # 트레이너 급여 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS trainer_salary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month TEXT NOT NULL,
            base_salary REAL DEFAULT 0,
            incentive REAL DEFAULT 0,
            tuition_fee REAL DEFAULT 0,  -- 수업료
            base_incentive_after_tax REAL,  -- 기본급+인센 3.3% 적용
            tuition_after_tax REAL,  -- 수업료 3.3% 적용
            incentive_payment_date DATE,
            tuition_payment_date DATE,
            total_salary REAL,
            total_after_tax REAL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES employees(id),
            UNIQUE(employee_id, year, month)
        )
    ''')

    # 인포 직원 급여 테이블
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS info_staff_salary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            year INTEGER NOT NULL,
            month TEXT NOT NULL,
            base_salary REAL DEFAULT 0,
            salary_after_tax REAL,  -- 3.3% 적용
            payment_date DATE,
            extra_pay REAL DEFAULT 0,  -- 부가급/인센티브
            extra_pay_note TEXT,  -- 부가급 비고 (ex: 식대)
            total_salary REAL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES employees(id),
            UNIQUE(employee_id, year, month)
        )
    ''')

    conn.commit()
    conn.close()
    print(f"✅ DB 생성 완료: {DB_PATH}")


def parse_date(value):
    """날짜 파싱"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        except:
            return None
    return None


def parse_number(value):
    """숫자 파싱 (수식 결과 포함)"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value.startswith('='):
            return 0  # 수식은 계산된 값으로 대체 필요
        try:
            return float(value.replace(',', ''))
        except:
            return 0
    return 0


def load_employees_from_excel():
    """인적사항2 시트에서 직원 정보 로드"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['인적사항2']

    employees = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, job_type, bank, account, resident_num, status = row[:6]
        if name and name.strip():
            employees.append({
                'name': name.strip(),
                'job_type': job_type.strip() if job_type else '',
                'bank': bank.strip() if bank else '',
                'account_number': str(account).strip() if account else '',
                'resident_number': str(resident_num).strip() if resident_num else '',
                'status': status.strip() if status else '근무'
            })

    wb.close()
    return employees


def load_trainer_salary_from_excel():
    """트레이너 시트에서 급여 정보 로드"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['트레이너']

    salary_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue

        # 급여월 파싱 (ex: "3월" -> 3)
        month_str = str(row[2]).replace('월', '').strip() if row[2] else ''

        salary_data.append({
            'name': str(name).strip(),
            'job_type': str(row[1]).strip() if row[1] else '트레이너',
            'month': month_str,
            'year': 2025,  # 기본값
            'base_salary': parse_number(row[3]),
            'incentive': parse_number(row[4]),
            'base_incentive_after_tax': parse_number(row[5]),
            'incentive_payment_date': parse_date(row[6]),
            'tuition_fee': parse_number(row[7]),
            'tuition_after_tax': parse_number(row[8]),
            'tuition_payment_date': parse_date(row[9]),
            'total_salary': parse_number(row[10]),
            'total_after_tax': parse_number(row[11]),
            'account_number': str(row[12]).strip() if row[12] else '',
            'bank': str(row[13]).strip() if row[13] else '',
            'resident_number': str(row[14]).strip() if row[14] else '',
            'notes': str(row[15]).strip() if row[15] else ''
        })

    wb.close()
    return salary_data


def load_info_salary_from_excel():
    """직원(인포) 시트에서 급여 정보 로드"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['직원(인포)']

    salary_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue

        month_str = str(row[2]).replace('월', '').strip() if row[2] else ''

        salary_data.append({
            'name': str(name).strip(),
            'job_type': str(row[1]).strip() if row[1] else '인포',
            'month': month_str,
            'year': 2025,
            'base_salary': parse_number(row[3]),
            'salary_after_tax': parse_number(row[4]),
            'payment_date': parse_date(row[5]),
            'extra_pay': parse_number(row[6]),
            'extra_pay_note': str(row[7]).strip() if row[7] else '',
            'total_salary': parse_number(row[8]),
            'account_number': str(row[9]).strip() if row[9] else '',
            'bank': str(row[10]).strip() if row[10] else '',
            'resident_number': str(row[11]).strip() if row[11] else ''
        })

    wb.close()
    return salary_data


def insert_employees(employees):
    """직원 정보 DB 입력"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    inserted = 0
    for emp in employees:
        try:
            cursor.execute('''
                INSERT OR REPLACE INTO employees
                (name, job_type, bank, account_number, resident_number, status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (emp['name'], emp['job_type'], emp['bank'],
                  emp['account_number'], emp['resident_number'], emp['status']))
            inserted += 1
        except Exception as e:
            print(f"  ⚠️ 직원 입력 실패: {emp['name']} - {e}")

    conn.commit()
    conn.close()
    print(f"✅ 직원 {inserted}명 입력 완료")
    return inserted


def get_employee_id(cursor, name):
    """직원 이름으로 ID 조회"""
    cursor.execute('SELECT id FROM employees WHERE name = ?', (name,))
    result = cursor.fetchone()
    return result[0] if result else None


def insert_trainer_salary(salary_data):
    """트레이너 급여 DB 입력"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    inserted = 0
    for sal in salary_data:
        emp_id = get_employee_id(cursor, sal['name'])

        # 직원이 없으면 추가
        if not emp_id:
            cursor.execute('''
                INSERT INTO employees (name, job_type, bank, account_number, resident_number)
                VALUES (?, ?, ?, ?, ?)
            ''', (sal['name'], sal['job_type'], sal['bank'],
                  sal['account_number'], sal['resident_number']))
            emp_id = cursor.lastrowid

        try:
            cursor.execute('''
                INSERT OR REPLACE INTO trainer_salary
                (employee_id, year, month, base_salary, incentive, tuition_fee,
                 base_incentive_after_tax, tuition_after_tax,
                 incentive_payment_date, tuition_payment_date,
                 total_salary, total_after_tax, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, sal['year'], sal['month'], sal['base_salary'],
                  sal['incentive'], sal['tuition_fee'], sal['base_incentive_after_tax'],
                  sal['tuition_after_tax'], sal['incentive_payment_date'],
                  sal['tuition_payment_date'], sal['total_salary'],
                  sal['total_after_tax'], sal['notes']))
            inserted += 1
        except Exception as e:
            print(f"  ⚠️ 트레이너 급여 입력 실패: {sal['name']} {sal['month']}월 - {e}")

    conn.commit()
    conn.close()
    print(f"✅ 트레이너 급여 {inserted}건 입력 완료")
    return inserted


def insert_info_salary(salary_data):
    """인포 직원 급여 DB 입력"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    inserted = 0
    for sal in salary_data:
        emp_id = get_employee_id(cursor, sal['name'])

        # 직원이 없으면 추가
        if not emp_id:
            cursor.execute('''
                INSERT INTO employees (name, job_type, bank, account_number, resident_number)
                VALUES (?, ?, ?, ?, ?)
            ''', (sal['name'], sal['job_type'], sal['bank'],
                  sal['account_number'], sal['resident_number']))
            emp_id = cursor.lastrowid

        try:
            cursor.execute('''
                INSERT OR REPLACE INTO info_staff_salary
                (employee_id, year, month, base_salary, salary_after_tax,
                 payment_date, extra_pay, extra_pay_note, total_salary)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, sal['year'], sal['month'], sal['base_salary'],
                  sal['salary_after_tax'], sal['payment_date'], sal['extra_pay'],
                  sal['extra_pay_note'], sal['total_salary']))
            inserted += 1
        except Exception as e:
            print(f"  ⚠️ 인포 급여 입력 실패: {sal['name']} {sal['month']}월 - {e}")

    conn.commit()
    conn.close()
    print(f"✅ 인포 급여 {inserted}건 입력 완료")
    return inserted


def show_summary():
    """DB 요약 출력"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    print("\n" + "="*50)
    print("📊 DB 요약")
    print("="*50)

    # 직원 수
    cursor.execute('SELECT job_type, status, COUNT(*) FROM employees GROUP BY job_type, status')
    print("\n👥 직원 현황:")
    for row in cursor.fetchall():
        print(f"  - {row[0]} ({row[1]}): {row[2]}명")

    # 트레이너 급여 현황
    cursor.execute('''
        SELECT e.name, COUNT(*), SUM(t.total_salary)
        FROM trainer_salary t
        JOIN employees e ON t.employee_id = e.id
        GROUP BY e.name
    ''')
    print("\n💰 트레이너 급여 현황:")
    for row in cursor.fetchall():
        total = row[2] if row[2] else 0
        print(f"  - {row[0]}: {row[1]}건, 총 {total:,.0f}원")

    # 인포 급여 현황
    cursor.execute('''
        SELECT e.name, COUNT(*), SUM(i.total_salary)
        FROM info_staff_salary i
        JOIN employees e ON i.employee_id = e.id
        GROUP BY e.name
    ''')
    print("\n💰 인포 급여 현황:")
    for row in cursor.fetchall():
        total = row[2] if row[2] else 0
        print(f"  - {row[0]}: {row[1]}건, 총 {total:,.0f}원")

    conn.close()


def main():
    print("="*50)
    print("🚀 급여 관리 DB 생성 시작")
    print("="*50)
    print(f"📁 엑셀 파일: {EXCEL_PATH}")
    print(f"📁 DB 파일: {DB_PATH}")

    # 1. DB 생성
    create_database()

    # 2. 직원 정보 로드 및 입력
    print("\n📥 직원 정보 로드 중...")
    employees = load_employees_from_excel()
    print(f"  - {len(employees)}명 로드됨")
    insert_employees(employees)

    # 3. 트레이너 급여 로드 및 입력
    print("\n📥 트레이너 급여 정보 로드 중...")
    trainer_salary = load_trainer_salary_from_excel()
    print(f"  - {len(trainer_salary)}건 로드됨")
    insert_trainer_salary(trainer_salary)

    # 4. 인포 급여 로드 및 입력
    print("\n📥 인포 급여 정보 로드 중...")
    info_salary = load_info_salary_from_excel()
    print(f"  - {len(info_salary)}건 로드됨")
    insert_info_salary(info_salary)

    # 5. 요약 출력
    show_summary()

    print("\n" + "="*50)
    print("✅ 급여 관리 DB 생성 완료!")
    print("="*50)


if __name__ == '__main__':
    main()
